"""
travel_seo_analyzer — streamlit_app.py
Interface Streamlit complète : scraping + 5 analyses SEO + export Excel.
"""
import io, os, shutil, datetime
import streamlit as st
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from scraper  import parse_sitemap, scrape_page, get_cache_status, SITEMAP_URL, CACHE_DIR
from analyzer import analyze_basic, analyze_offers, analyze_spec, analyze_google_ai, extract_content

# ─────────────────────────────────────────────────────────────────────────────
# Config (doit être le premier appel Streamlit)
# ─────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="travel_seo_analyzer",
    page_icon="⬢",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""<style>
[data-testid="stAppViewContainer"]  { background:#f4f6f9; }
[data-testid="stSidebar"]           { background:#1e3a6e; }
[data-testid="stSidebar"] *         { color:#d0def7 !important; }
[data-testid="stSidebar"] h1,
[data-testid="stSidebar"] h2,
[data-testid="stSidebar"] h3        { color:#ffffff !important; }
.block-container                    { padding-top:1.5rem; }
.stTabs [data-baseweb="tab-list"]   { gap:4px; background:#fff; border-radius:8px;
                                      padding:4px; border:1px solid #d5dae5; }
.stTabs [data-baseweb="tab"]        { border-radius:6px; padding:6px 16px;
                                      font-size:13px; font-weight:500; color:#7a84a0; }
.stTabs [aria-selected="true"]      { background:#1e3a6e !important; color:#fff !important; }
div[data-testid="metric-container"] { background:#fff; border:1px solid #d5dae5;
                                      border-radius:10px; padding:16px; }
.stButton > button                  { background:#1e3a6e; color:#fff;
                                      border:none; border-radius:6px; font-weight:500; }
.stButton > button:hover            { background:#2a4f96; color:#fff; }
</style>""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# Session state
# ─────────────────────────────────────────────────────────────────────────────
KEYS = ["urls","data_basic","data_offers","data_spec","data_google_ai","data_content"]
for k in KEYS:
    if k not in st.session_state:
        st.session_state[k] = None

# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────
def today() -> str:
    return datetime.date.today().isoformat()

def status_icon(v) -> str:
    if v is None: return "—"
    return "✅" if v else "❌"

def needs_cache() -> bool:
    """Affiche un message si pas d'URLs et retourne True."""
    if not st.session_state.urls:
        st.info("👈 **Étape 1** : charge les URLs via la sidebar\n\n👈 **Étape 2** : lance le scraping")
        return True
    return False

def run_analysis(key: str, fn, label: str):
    """Lance l'analyse si non encore effectuée."""
    if st.session_state[key] is None and st.session_state.urls:
        with st.spinner(f"{label}…"):
            try:
                st.session_state[key] = fn(st.session_state.urls)
            except Exception as e:
                st.error(f"Erreur d'analyse : {e}")
    return st.session_state[key]

def to_excel(df: pd.DataFrame, sheet_name: str) -> bytes:
    buf = io.BytesIO()
    wb  = Workbook()
    ws  = wb.active
    if ws is None:
        ws = wb.create_sheet()
    ws.title = sheet_name[:31]

    HDR_FILL = PatternFill("solid", fgColor="1E3A6E")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALT_FILL  = PatternFill("solid", fgColor="EEF3FB")
    thin      = Side(style="thin", color="D5DAE5")
    BORDER    = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ci, h in enumerate(df.columns, 1):
        c = ws.cell(row=1, column=ci, value=str(h))
        c.fill, c.font, c.alignment, c.border = HDR_FILL, HDR_FONT, HDR_ALIGN, BORDER

    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(vertical="top")
            c.border = BORDER
            if ri % 2 == 0:
                c.fill = ALT_FILL

    for ci, col in enumerate(df.columns, 1):
        w = max(len(str(col)), df[col].astype(str).str.len().max() if len(df) else 0)
        ws.column_dimensions[get_column_letter(ci)].width = min(60, max(10, w + 2))

    ws.freeze_panes = "A2"
    wb.save(buf)
    return buf.getvalue()

def _clean_df(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalise les types du DataFrame avant affichage ou export.
    - Colonnes entières/mixtes → str pour Streamlit/Arrow
    - None → chaîne vide dans l'affichage
    """
    df = df.copy()
    for col in df.columns:
        if col in ("URL",):
            continue
        # Si colonne mixte (int + str ou None + int) → tout en str sauf si purement numérique
        if df[col].dtype == object:
            df[col] = df[col].apply(lambda v: "" if v is None else str(v) if not isinstance(v, str) else v)
    return df

def btn_excel(df: pd.DataFrame, filename: str, sheet: str) -> None:
    st.download_button(
        label="📥 Télécharger Excel",
        data=to_excel(df, sheet),
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=filename,
    )

def show_table(df: pd.DataFrame, total: int) -> None:
    st.caption(f"{len(df):,} / {total:,} pages affichées")
    # Normaliser les types pour éviter l'erreur Arrow (mixed int/str)
    for col in df.columns:
        if col not in ("URL",):
            try:
                df[col] = df[col].where(df[col].notna(), other=None)
            except Exception:
                pass
    df = _clean_df(df)
    st.dataframe(
        df,
        use_container_width=True,
        height=450,
        column_config={"URL": st.column_config.LinkColumn()},
    )

def filter_df(df: pd.DataFrame, q: str, col_issues: str | None = None,
              col_filter: str | None = None, filter_val: str = "Tout") -> pd.DataFrame:
    if q:
        df = df[df["URL"].str.contains(q, case=False, na=False)]
    if col_issues and filter_val == "Avec problèmes":
        df = df[df[col_issues].astype(str).str.strip() != ""]
    if col_issues and filter_val in ("Sans problème", "Conformes", "Optimisées"):
        df = df[df[col_issues].astype(str).str.strip() == ""]
    return df

# ─────────────────────────────────────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⬢ travel_seo_analyzer")
    st.markdown("---")

    # Statut cache
    cache = get_cache_status()
    if cache["exists"]:
        dt = datetime.datetime.fromisoformat(cache["date"]).strftime("%d/%m %H:%M")
        st.success(f"Cache actif · {cache['count']} pages · {dt}")
    else:
        st.warning("Pas de cache — lancez le scraping")

    st.markdown("---")
    st.markdown("### 1 · Sitemap")
    sitemap_url = st.text_input("URL du sitemap", value=SITEMAP_URL,
                                label_visibility="collapsed")

    if st.button("📋 Charger les URLs", use_container_width=True):
        with st.spinner("Parsing du sitemap…"):
            try:
                st.session_state.urls = parse_sitemap(sitemap_url)
                if st.session_state.urls:
                    st.success(f"✅ {len(st.session_state.urls)} URLs")
                else:
                    st.error("0 URL trouvée — vérifiez l'URL du sitemap")
            except Exception as e:
                st.error(f"Erreur : {e}")

    if st.session_state.urls:
        st.info(f"**{len(st.session_state.urls):,}** URLs chargées")

    st.markdown("---")
    st.markdown("### 2 · Scraping")
    force = st.checkbox("Forcer le re-téléchargement (ignore le cache)")

    if st.button("▶ Lancer le scraping", type="primary", use_container_width=True):
        # ── Auto-chargement des URLs si nécessaire ────────────────────────────
        if not st.session_state.urls:
            with st.spinner("Chargement du sitemap…"):
                try:
                    st.session_state.urls = parse_sitemap(sitemap_url)
                    if not st.session_state.urls:
                        st.error("0 URL trouvée dans le sitemap — scraping annulé.")
                        st.stop()
                    st.success(f"✅ {len(st.session_state.urls)} URLs chargées automatiquement")
                except Exception as e:
                    st.error(f"Impossible de charger le sitemap : {e}")
                    st.stop()

        urls  = st.session_state.urls
        total = len(urls)
        errs  = []

        bar  = st.progress(0.0, text=f"0 / {total}")
        info = st.empty()

        # ── Pré-chargement de l'index UNE seule fois ──────────────────────────
        # Evite les lectures/écritures répétées sur le disque à chaque URL,
        # ce qui provoquait le double-scraping observé en logs.
        from scraper import _slug, _decode, _save_index, HEADERS, CACHE_DIR
        import os, time, requests as _req
        from datetime import datetime as _dt

        os.makedirs(CACHE_DIR, exist_ok=True)
        idx = load_cache_index()   # lecture unique

        for i, url in enumerate(urls):
            filename   = _slug(url)
            cache_path = os.path.join(CACHE_DIR, filename)

            # Déjà en cache et pas de force → skip sans re-requête
            if not force and url in idx and os.path.exists(cache_path):
                icon = "📁"
            else:
                time.sleep(0.3)
                try:
                    resp   = _req.get(url, headers=HEADERS, timeout=30, allow_redirects=True)
                    status = resp.status_code
                    html   = f"<!-- HTTP_{status} -->" if status >= 400 else _decode(resp.content)
                    if status >= 400:
                        errs.append(f"HTTP {status} · {url}")
                    icon = "⚠" if status >= 400 else "✓"
                except Exception as exc:
                    html, status = f"<!-- SCRAPE_ERROR: {exc} -->", 0
                    errs.append(f"Erreur · {url} → {exc}")
                    icon = "✗"

                with open(cache_path, "w", encoding="utf-8") as f:
                    f.write(html)

                idx[url] = {"file": filename,
                            "scraped_at": _dt.now().isoformat(),
                            "status": status}

            pct   = (i + 1) / total
            short = url.replace("https://voyage.showroomprive.com", "")[:45]
            bar.progress(pct, text=f"{icon} {i+1}/{total} — {short}")
            info.caption(url)

        # ── Sauvegarde de l'index UNE seule fois à la fin ────────────────────
        _save_index(idx)

        bar.progress(1.0, text="✅ Terminé")
        info.empty()

        ok_count = total - len([e for e in errs if not e.startswith("HTTP 401")])
        if errs:
            nb_401 = sum(1 for e in errs if "401" in e)
            nb_other = len(errs) - nb_401
            msg = []
            if nb_401:   msg.append(f"{nb_401} pages en 401 (accès refusé)")
            if nb_other: msg.append(f"{nb_other} autre(s) erreur(s)")
            st.warning("⚠ " + " · ".join(msg))
        else:
            st.success(f"✅ {total} pages scrappées sans erreur")

        # Invalider les analyses pour forcer recalcul
        for k in ["data_basic","data_offers","data_spec","data_google_ai","data_content"]:
            st.session_state[k] = None
        st.rerun()

    st.markdown("---")
    if st.button("🗑 Vider le cache", use_container_width=True):
        if os.path.exists(CACHE_DIR):
            shutil.rmtree(CACHE_DIR)
        for k in KEYS:
            st.session_state[k] = None
        st.success("Cache vidé")
        st.rerun()

# ─────────────────────────────────────────────────────────────────────────────
# En-tête principal
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("# ⬢ travel_seo_analyzer")
st.caption("voyage.showroomprive.com · Analyse SEO")

# Métriques globales
c1, c2, c3, c4, c5 = st.columns(5)
with c1: st.metric("URLs sitemap",  len(st.session_state.urls) if st.session_state.urls else "—")
with c2: st.metric("Problèmes SEO", sum(1 for r in (st.session_state.data_basic or []) if r.get("issues")) or "—")
with c3: st.metric("Sans offres",   sum(1 for r in (st.session_state.data_offers or []) if not r.get("has_offers") and not r.get("error")) or "—")
with c4: st.metric("Avec offres",   sum(1 for r in (st.session_state.data_offers or []) if r.get("has_offers")) or "—")
with c5: st.metric("Erreurs HTTP",  sum(1 for r in (st.session_state.data_basic or []) if r.get("error")) or "—")

st.markdown("---")

# ─────────────────────────────────────────────────────────────────────────────
# Onglets
# ─────────────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "🔍 SEO basique",
    "🏷 Offres & Accès HTTP",
    "📝 Contenu éditorial",
    "⚡ Conformité specs",
    "🤖 Guide Google AIO",
])

# ── Tab 1 : SEO basique ───────────────────────────────────────────────────────
with tab1:
    if not needs_cache():
        c_l, c_r = st.columns([4,1])
        with c_r:
            if st.button("↻ Relancer", key="rl_basic"):
                st.session_state.data_basic = None
        data = run_analysis("data_basic", analyze_basic, "Analyse SEO basique")
        if data:
            df = pd.DataFrame([{
                "URL":         r["url"],
                "Statut":      str(str(r.get("status","") or "") or ""),
                "H1":          status_icon(r.get("h1_ok")),
                "H1 (valeur)": r.get("h1","")[:80],
                "Title":       status_icon(r.get("title_ok")),
                "Meta desc":   status_icon(r.get("meta_ok")),
                "Nb H1":       r.get("h1_count",""),
                "Problèmes":   " | ".join(r.get("issues",[])),
                "Erreur HTTP": r.get("error") or "",
            } for r in data])
            q1, q2 = st.columns([3,2])
            with q1: q   = st.text_input("Filtrer par URL", key="q_basic", placeholder="/spa, /ski…")
            with q2: flt = st.selectbox("Afficher", ["Tout","Avec problèmes","Sans problème"], key="f_basic")
            df = filter_df(df, q, "Problèmes", None, flt)
            if flt == "Avec problèmes": df = df[df["Problèmes"].str.strip() != ""]
            if flt == "Sans problème":  df = df[df["Problèmes"].str.strip() == ""]
            show_table(df, len(data))
            btn_excel(df, f"seo_basic_{today()}.xlsx", "SEO Basique")

# ── Tab 2 : Offres & Accès HTTP ───────────────────────────────────────────────
with tab2:
    if not needs_cache():
        c_l, c_r = st.columns([4,1])
        with c_r:
            if st.button("↻ Relancer", key="rl_offers"):
                st.session_state.data_offers = None
        data = run_analysis("data_offers", analyze_offers, "Analyse offres")
        if data:
            df = pd.DataFrame([{
                "URL":         r["url"],
                "Statut HTTP": str(str(r.get("status","") or "") or ""),
                "Offres":      "✅ Oui" if r.get("has_offers") else ("⚠ Erreur" if r.get("error") else "❌ Non"),
                "Nb offres":   r.get("count"),  # None si absent — évite le mixed-type Arrow
                "Erreur":      r.get("error") or "",
            } for r in data])
            q1, q2 = st.columns([3,2])
            with q1: q   = st.text_input("Filtrer par URL", key="q_offers")
            with q2: flt = st.selectbox("Afficher", ["Tout","Sans offres","Avec offres","Erreurs HTTP"], key="f_offers")
            if q: df = df[df["URL"].str.contains(q, case=False, na=False)]
            if flt == "Sans offres":  df = df[df["Offres"] == "❌ Non"]
            if flt == "Avec offres":  df = df[df["Offres"] == "✅ Oui"]
            if flt == "Erreurs HTTP": df = df[df["Erreur"] != ""]
            show_table(df, len(data))
            btn_excel(df, f"seo_offers_{today()}.xlsx", "Offres & Accès HTTP")

# ── Tab 3 : Contenu éditorial ─────────────────────────────────────────────────
with tab3:
    if not needs_cache():
        c_l, c_r = st.columns([4,1])
        with c_r:
            if st.button("↻ Relancer", key="rl_content"):
                st.session_state.data_content = None
        data = run_analysis("data_content", extract_content, "Extraction contenu éditorial")
        if data:
            df = pd.DataFrame([{
                "URL":               r["url"],
                "Statut HTTP":       str(str(r.get("status","") or "") or ""),
                "H1":                r.get("h1","")[:120],
                "Meta Title":        r.get("meta_title","")[:120],
                "Meta Description":  r.get("meta_desc","")[:160],
                "header-intro":      r.get("header_intro","")[:200],
                "bloc-edito-content":r.get("bloc_edito","")[:200],
                "cpt-faq":           r.get("cpt_faq","")[:200],
                "Erreur":            r.get("error") or "",
            } for r in data])
            q = st.text_input("Filtrer par URL", key="q_content")
            if q: df = df[df["URL"].str.contains(q, case=False, na=False)]
            show_table(df, len(data))
            btn_excel(df, f"seo_content_{today()}.xlsx", "Contenu éditorial")

# ── Tab 4 : Conformité specs ──────────────────────────────────────────────────
with tab4:
    if not needs_cache():
        c_l, c_r = st.columns([4,1])
        with c_r:
            if st.button("↻ Relancer", key="rl_spec"):
                st.session_state.data_spec = None
        data = run_analysis("data_spec", analyze_spec, "Conformité specs")
        if data:
            rows = []
            for r in data:
                row: dict = {"URL":r["url"], "Statut":str(str(r.get("status","") or "") or ""), "Score %":r.get("pct","")}
                for k,v in r.get("checks",{}).items():
                    row[k] = status_icon(v)
                row["Problèmes"] = " | ".join(r.get("issues",[]))
                row["Erreur"]    = r.get("error") or ""
                rows.append(row)
            df = pd.DataFrame(rows)
            q1, q2 = st.columns([3,2])
            with q1: q   = st.text_input("Filtrer par URL", key="q_spec")
            with q2: flt = st.selectbox("Afficher", ["Tout","Avec problèmes","Conformes"], key="f_spec")
            if q: df = df[df["URL"].str.contains(q, case=False, na=False)]
            if flt == "Avec problèmes": df = df[df["Problèmes"].str.strip() != ""]
            if flt == "Conformes":      df = df[df["Problèmes"].str.strip() == ""]
            show_table(df, len(data))
            btn_excel(df, f"seo_spec_{today()}.xlsx", "Conformité Specs")

# ── Tab 5 : Guide Google AIO ──────────────────────────────────────────────────
with tab5:
    if not needs_cache():
        c_l, c_r = st.columns([4,1])
        with c_r:
            if st.button("↻ Relancer", key="rl_gai"):
                st.session_state.data_google_ai = None
        data = run_analysis("data_google_ai", analyze_google_ai, "Guide Google AIO")
        if data:
            rows = []
            for r in data:
                row = {"URL":r["url"], "Statut":str(str(r.get("status","") or "") or ""),
                       "Score %":r.get("pct",""), "Nb mots":r.get("words","")}
                for k,v in r.get("checks",{}).items():
                    row[k] = status_icon(v)
                row["Problèmes"] = " | ".join(r.get("issues",[]))
                row["Erreur"]    = r.get("error") or ""
                rows.append(row)
            df = pd.DataFrame(rows)
            q1, q2 = st.columns([3,2])
            with q1: q   = st.text_input("Filtrer par URL", key="q_gai")
            with q2: flt = st.selectbox("Afficher", ["Tout","Avec problèmes","Optimisées"], key="f_gai")
            if q: df = df[df["URL"].str.contains(q, case=False, na=False)]
            if flt == "Avec problèmes": df = df[df["Problèmes"].str.strip() != ""]
            if flt == "Optimisées":     df = df[df["Problèmes"].str.strip() == ""]
            show_table(df, len(data))
            btn_excel(df, f"seo_google_ai_{today()}.xlsx", "Guide Google AIO")
